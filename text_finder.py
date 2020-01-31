import openpyxl
from collections import Counter
import string
from openpyxl import load_workbook


def data_from_market():
    wb = load_workbook('test.xlsx')
    worksheet = wb['APR_Market']
    headers = []
    for colnum in range(1, worksheet.max_column + 1):
        headers.append(worksheet.cell(column=colnum, row=10).value)
    print (headers)
    data_from_market = []
    for row in range(11, worksheet.max_row +1):
        line = {}
        for header in headers:
            cell_value = worksheet.cell(column=headers.index(header)+1,row=row).value
            if type(cell_value) == 'str':
                cell_value = cell_value.strip()
            elif type(cell_value) == 'int':
                cell_value = str(cell_value)
            elif cell_value is None:
                pass
            line[header] = cell_value
            data_from_market.append(line)
    return data_from_market

print(data_from_market[0:6])
#print(data_from_market)
#print(data_from_market[0])


def check_for_valid_data(data_from_market):
    is_valid_data = (
                    str(line["Способ показа"]) != "Экран"
                    and line["OTS"]) <= 0
                    and line["Выходов в сутки"] <= 0)
    
    return is_valid_data

#def check_for_A51(data_from_market):
    #bb_A51 = ['MSBB00263A', 'MSBB00313A', 'MSBB12969A', 'MSBB13526A']
    #is_bb_A51 = (line["GID"][:10] in  bb_A51)
    #return is_bb_A51   

def check_for_double_bb(data_from_market):
    double_bb = [MSBB04620A, MSBB15852A ]
    is_bb_A51 = bool((line["GID"][:10] in double_bb ))
    return is_double_bb


def check_for_side_A1(data_from_market):
    is_side_A1 = bool(
                ("из цента"  in line['Адрес поверхности'])
                 or ("до Х" in line['Адрес поверхности']))
    return is_side_A1

#def check_for_side_A51(data_from_market):
    #is_side_A51 = bool(
                #("в центр"  in line['Адрес поверхности'])
                 #or ("после Х " in line['Адрес поверхности']))
             
   # return is_side_A51

def billboards_data(data_from_market):
    if check_valid_data(data_from_market):
        raise ValueError("Не хватает данных")
   
    billboards_data = {}
    for line in data_from_market:
    if line["GID"][:10] not in billboards_data.keys():
        if check_for_double_bb(data_from_market):
            if f'{line["GID"][:10]}' not in billboards_data.keys():
                billboards_data.update({f'{line["GID"]}':[ float(line["OTS"]),  float(line["Выходов в сутки"]), str(line['Адрес поверхности'])] })
            else:
                
                #добавить к прошлому номеру поверхности, если совпало направление
                    if check_for_side_A1() == check_for_side_A1():
                # billboards_data[f'{line["GID"][:10]}51'][1] += float(line["Выходов в сутки"])
                    else:
                        billboards_data.update({f'{line["GID"]}':[ float(line["OTS"]),  float(line["Выходов в сутки"]), str(line['Адрес поверхности'])] })
     #добавить к прошлому номеру поверхности
    else:

        billboards_data[f'{line["GID"][:10]}'][1] += float(line["Выходов в сутки"]) 
    

        

